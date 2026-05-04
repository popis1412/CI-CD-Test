import openpyxl
import os
import sys

# =========================
# 사용자 설정
# =========================
header_row = 9

def normalize(text):
    return str(text).replace(" ", "").strip().lower() if text else ""

# 상태 정의
status_keys = ["Pass", "Fail", "Not Test", "Blocked", "N/A"]

normalized_status_map = {
    "pass": "Pass",
    "fail": "Fail",
    "nottest": "Not Test",
    "blocked": "Blocked",
    "n/a": "N/A",
    "na": "N/A"
}

BASE_PATH = r"C:\QA\CI-CD-Test"

# =========================
# 경로
# =========================
if len(sys.argv) >= 2:
    base_path = sys.argv[1]
else:
    base_path = BASE_PATH

excel_path = os.path.join(base_path, 'Tests', 'QA_Test.xlsx')
output_dir = os.path.join(base_path, 'Test Results')
output_html = os.path.join(output_dir, 'qa_report.html')

os.makedirs(output_dir, exist_ok=True)

wb = openpyxl.load_workbook(excel_path, data_only=True)
target_sheets = [s for s in wb.sheetnames if s.startswith("TC_")]

# =========================
# 통계 초기화
# =========================
overall = {k: 0 for k in status_keys}
overall["Total"] = 0

sheet_stats = {}
category_stats = {}
defects = []

# =========================
# 데이터 처리
# =========================
for sheet_name in target_sheets:
    ws = wb[sheet_name]

    sheet_stats[sheet_name] = {k: 0 for k in status_keys}
    sheet_stats[sheet_name]["Total"] = 0

    # 헤더 찾기
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            header_map[normalize(val)] = col

    result_col = header_map.get("수행결과")
    cat1_col = header_map.get("대분류")
    cat2_col = header_map.get("중분류")
    action_col = header_map.get("테스트액션")

    if not result_col:
        continue

    for row in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(row=row, column=result_col).value
        status_norm = normalize(raw)

        # ✔ 값이 있으면 무조건 Total
        if status_norm:
            overall["Total"] += 1
            sheet_stats[sheet_name]["Total"] += 1

            # ✔ 상태 매핑
            mapped = normalized_status_map.get(status_norm)

            if mapped:
                overall[mapped] += 1
                sheet_stats[sheet_name][mapped] += 1

                cat1 = ws.cell(row=row, column=cat1_col).value if cat1_col else "미분류"
                cat2 = ws.cell(row=row, column=cat2_col).value if cat2_col else "미분류"
                action = ws.cell(row=row, column=action_col).value if action_col else ""

                cat1 = cat1 if cat1 else "미분류"
                cat2 = cat2 if cat2 else "미분류"

                key = (cat1, cat2, mapped)
                category_stats[key] = category_stats.get(key, 0) + 1

                if mapped == "Fail":
                    defects.append({
                        "Sheet": sheet_name,
                        "Cat1": cat1,
                        "Cat2": cat2,
                        "Row": row,
                        "Action": action
                    })

# =========================
# 퍼센트
# =========================
def pct(v, t):
    return f"{(v/t*100):.1f}%" if t else "0.0%"

# =========================
# HTML 생성
# =========================
# =========================
# HTML 생성 (디자인 적용)
# =========================
html = f'''
<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>QA 정량 분석 리포트</title>

<style>
body {{
    font-family: 'Segoe UI', 'Malgun Gothic', sans-serif;
    margin: 40px;
    color: #333;
    background-color: #f9f9f9;
}}

.container {{
    background-color: white;
    padding: 30px;
    border-radius: 10px;
    box-shadow: 0 0 20px rgba(0,0,0,0.05);
}}

h1 {{
    color: #1a2a6c;
    border-bottom: 4px solid #1a2a6c;
    padding-bottom: 15px;
    text-align: center;
}}

h2 {{
    color: #2c3e50;
    margin-top: 50px;
    border-left: 5px solid #1a2a6c;
    padding-left: 15px;
}}

h3 {{
    color: #555;
    margin-top: 30px;
}}

table {{
    border-collapse: collapse;
    width: 100%;
    margin-bottom: 30px;
    border-radius: 8px;
    overflow: hidden;
}}

th, td {{
    border: 1px solid #eee;
    padding: 14px;
    text-align: center;
}}

th {{
    background-color: #f1f3f5;
    font-weight: 700;
    color: #444;
}}

tr:hover {{
    background-color: #f8f9fa;
}}

.pass {{ color: #2ecc71; font-weight: bold; }}
.fail {{ color: #e74c3c; font-weight: bold; }}
.not-test {{ color: #f39c12; font-weight: bold; }}
.blocked {{ color: #9b59b6; font-weight: bold; }}

.total-row {{
    background-color: #f1f5f8;
    font-weight: bold;
}}

.defect-table {{ text-align: left; }}
.center {{ text-align: center; }}
</style>
</head>

<body>
<div class="container">

<h1>📋 QA 테스트 결과 정량 분석 보고서</h1>

<!-- ========================= -->
<!-- 0. Overall -->
<!-- ========================= -->
<h2>0. 테스트 요약 (Overall)</h2>
<table>
<tr><th>항목</th><th>카운트</th><th>비율</th></tr>

<tr class="total-row">
<td>Total Test Cases</td>
<td>{overall["Total"]}</td>
<td>100.0%</td>
</tr>

<tr>
<td><span class="pass">Pass</span></td>
<td>{overall["Pass"]}</td>
<td>{pct(overall["Pass"], overall["Total"])}</td>
</tr>

<tr>
<td><span class="fail">Fail</span></td>
<td>{overall["Fail"]}</td>
<td>{pct(overall["Fail"], overall["Total"])}</td>
</tr>

<tr>
<td><span class="not-test">Not Test</span></td>
<td>{overall["Not Test"]}</td>
<td>{pct(overall["Not Test"], overall["Total"])}</td>
</tr>

<tr>
<td><span class="blocked">Blocked</span></td>
<td>{overall["Blocked"]}</td>
<td>{pct(overall["Blocked"], overall["Total"])}</td>
</tr>

<tr>
<td>N/A</td>
<td>{overall["N/A"]}</td>
<td>{pct(overall["N/A"], overall["Total"])}</td>
</tr>

</table>

<!-- ========================= -->
<!-- 1. 카테고리 -->
<!-- ========================= -->
<h2>1. 대분류 / 중분류 별 테스트 현황</h2>
<table>
<tr><th>대분류</th><th>중분류</th><th>현황</th><th>개수</th></tr>
'''

# 카테고리 출력
for (cat1, cat2, status), count in category_stats.items():
    cls = status.lower().replace(" ", "-")
    html += f'<tr><td>{cat1}</td><td>{cat2}</td><td><span class="{cls}">{status}</span></td><td>{count}</td></tr>'

html += '</table>'


# =========================
# 2. 시나리오 실패율
# =========================
html += '<h2>2. 시나리오별 실패 비율</h2>'

for sheet, stats in sheet_stats.items():
    fail_rate = pct(stats["Fail"], stats["Total"])
    html += f'''
    <h3>📁 {sheet}</h3>
    <table>
    <tr><th>시나리오</th><th>실패 비율</th></tr>
    <tr>
        <td>{sheet}</td>
        <td class="fail">{fail_rate}</td>
    </tr>
    </table>
    '''

# =========================
# 3. 결함 리스트
# =========================
html += '''
<h2>3. 발견된 결함 상세 (Defect List)</h2>
<table class="defect-table">
<tr>
<th>No</th><th>시나리오</th><th>대분류</th>
<th>중분류</th><th>행</th><th>테스트 액션</th>
</tr>
'''

for i, d in enumerate(defects, 1):
    html += f'''
    <tr>
        <td class="center">{i}</td>
        <td class="center">{d["Sheet"]}</td>
        <td class="center">{d["Cat1"]}</td>
        <td class="center">{d["Cat2"]}</td>
        <td class="center">{d["Row"]}</td>
        <td>{d["Action"]}</td>
    </tr>
    '''

html += '''
</table>
</div>
</body>
</html>
'''

# =========================
# 파일 저장
# =========================
with open(output_html, "w", encoding="utf-8") as f:
    f.write(html)

print(f"리포트 생성 완료: {output_html}")
print(f"[FAIL_COUNT] {overall['Fail']}")

# 파일 저장
with open(output_html, "w", encoding="utf-8") as f:
    f.write(html)

print(f"리포트 생성 완료: {output_html}")
print(f"[FAIL_COUNT] {overall['Fail']}")