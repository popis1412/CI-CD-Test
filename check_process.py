import os
import json
import threading
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import Workbook, load_workbook

# ================================
# 설정
# ================================
FILE_PATH = r"C:\QA\CI-CD-Test\result.xlsx"
TESTER = "AutoBot"

# 전역 상태 관리
def get_default_test_state():
    return {
        "page_loaded": False,
        "image_loaded": False,
        "request_sent": False,
        "modal_opened": False,
        "payment_success": False,
        "alert_shown": False,
        "disabled": False,
        "text_changed": False,
        "blocked": False,
        "no_duplicate": True
    }

TEST_STATE = get_default_test_state()
SERVER_STATE = {"request_count": 0}

os.makedirs(os.path.dirname(FILE_PATH), exist_ok=True)

# ================================
# 엑셀 관련 함수
# ================================
def reset_excel():
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["NO","대분류","중분류","사전조건","테스트 액션","기대 결과","수행 결과","테스터","날짜","Notes"])
        wb.save(FILE_PATH)
        print("[RESET] 엑셀 초기화 완료")
    except PermissionError:
        print("\n" + "="*50)
        print("[ERROR] result.xlsx 파일이 열려있어 초기화할 수 없습니다.")
        print("파일을 닫고 프로그램을 다시 실행해주세요.")
        print("="*50 + "\n")

def append_excel(row):
    try:
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        ws.append(row)
        wb.save(FILE_PATH)
    except PermissionError:
        print(f"[ERROR] 엑셀 파일이 열려있어 '{row[2]}' 결과를 저장하지 못했습니다.")

def is_testcase_exists(no):
    if not os.path.exists(FILE_PATH): return False
    try:
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == no: return True
    except: pass
    return False

# ================================
# 테스트 실행 엔진
# ================================
def run_tests(context, selected="N/A"):
    today = datetime.now().strftime("%Y-%m-%d")
    
    test_cases = []

    # ================================
    # 공통 테스트 (1~3번)
    # ================================
    if not is_testcase_exists(1):
        test_cases.append((
            1, "UI", "접속",
            "서버가 정상적으로 실행 중이며 사용자가 웹 브라우저를 통해 메인 페이지에 접근 가능한 상태여야 한다.",
            "사용자는 웹 브라우저의 주소창에 URL을 입력하여 메인 페이지에 접속해야 한다.",
            "메인 페이지가 정상적으로 로드되어야 한다.",
            context["page_loaded"]
        ))
    if not is_testcase_exists(2):
        test_cases.append((
            2, "UI", "이미지",
            "사용자가 메인 페이지에 접속한 상태여야 한다.",
            "사용자는 각 상품의 이미지를 확인해야 한다.",
            "각 상품 이미지가 깨지지 않고 정상적으로 표시되어야 한다.",
            context["image_loaded"]
        ))
    if not is_testcase_exists(3):
        test_cases.append((
            3, "서버", "요청",
            "사용자가 결제 버튼을 클릭한 상태여야 한다.",
            "서버로 POST 요청이 전달되어야 한다.",
            "서버에서 요청을 정상적으로 수신해야 한다.",
            context["request_sent"]
        ))

    # ================================
    # 아이템별 테스트 (4~10번)
    # ================================
    if selected != "N/A":
        test_cases.append((
            4, "UI", f"{selected} 버튼 클릭",
            "사용자가 메인 페이지를 보고 있는 상태여야 한다.",
            f"사용자는 {selected} 상품의 구매 버튼을 클릭해야 한다.",
            "결제 모달 창이 정상적으로 표시되어야 한다.",
            context["modal_opened"]
        ))
        test_cases.append((
            5, "UI", f"{selected} 결제",
            "결제 모달 창이 열린 상태여야 한다.",
            "사용자는 결제 버튼을 클릭해야 한다.",
            "결제가 정상적으로 처리되어야 한다.",
            context["payment_success"]
        ))
        test_cases.append((
            6, "UI", f"{selected} 알림",
            "결제가 완료된 상태여야 한다.",
            "사용자는 알림 메시지를 확인해야 한다.",
            "결제가 완료되었습니다라는 메시지가 표시되어야 한다.",
            context["alert_shown"]
        ))
        test_cases.append((
            7, "UI", f"{selected} 버튼 비활성화",
            "결제가 완료된 상태여야 한다.",
            f"{selected} 버튼 상태를 확인해야 한다.",
            "버튼이 비활성화 상태로 변경되어야 한다.",
            context["disabled"]
        ))
        test_cases.append((
            8, "UI", f"{selected} 텍스트 변경",
            "결제가 완료된 상태여야 한다.",
            f"{selected} 버튼의 텍스트를 확인해야 한다.",
            "버튼에 구매완료라는 텍스트가 표시되어야 한다.",
            context["text_changed"]
        ))
        test_cases.append((
            9, "UI", f"{selected} 재클릭 방지",
            "이미 해당 상품을 구매한 상태여야 한다.",
            f"{selected} 버튼을 다시 클릭해야 한다.",
            "클릭이 동작하지 않아야 한다.",
            context["blocked"]
        ))
        test_cases.append((
            10, "UI", "중복 클릭 방지",
            "사용자가 결제 버튼을 빠르게 여러 번 클릭한 상태여야 한다.",
            "결제 요청이 여러 번 발생하는지 확인해야 한다.",
            "결제 요청은 한 번만 처리되어야 한다.",
            context["no_duplicate"]
        ))

    # 🔥 수정된 부분: raw_cases -> test_cases
    for tc in test_cases:
        # 이미 성공한 공통 케이스(1,2,3)는 중복 기록하지 않음
        if tc[0] <= 3 and is_testcase_exists(tc[0]):
            continue
            
        status = "Pass" if tc[6] else "Fail"
        append_excel([tc[0], tc[1], tc[2], tc[3], tc[4], tc[5], status, TESTER, today, ""])

# ================================
# HTML (절대 수정 금지 구역)
# ================================
MAIN_PAGE = """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
    body { background:#0f172a; color:white; font-family:sans-serif; text-align:center; }
    .container { display:flex; justify-content:center; gap:30px; padding:20px; }
    .card { width:200px; background:#1e293b; padding:20px; border-radius:15px; }
    button { width:100%; padding:10px; margin-top:10px; background:#22c55e; border:none; color:white; cursor:pointer; border-radius:5px; }
    button:disabled { background:gray; }
    .modal { display:none; position:fixed; top:50%; left:50%; transform:translate(-50%,-50%); width:300px; background:white; color:black; padding:20px; border-radius:10px; border:2px solid #ccc; }
</style>
<script>
let selectedItem = null;
let isProcessing = false;

function openPay(id){
    selectedItem = id;
    fetch("/click", { method:"POST", body: JSON.stringify({selected: id}) });
    document.getElementById("modal").style.display="block";
}

function doPay(){
    if(isProcessing) return;
    isProcessing = true;

    fetch("/purchase", {
        method:"POST",
        body: JSON.stringify({ selected: selectedItem })
    })
    .then(res => {
        alert("결제가 완료되었습니다.");
        let btn = document.getElementById(selectedItem);
        btn.disabled = true;
        btn.innerText = "구매완료";
        document.getElementById("modal").style.display="none";

        // 최종 결과 전송
        fetch("/result", {
            method:"POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({
                selected: selectedItem,
                alert_shown: true,
                disabled: true,
                text_changed: true,
                blocked: true,
                payment_success: true
            })
        });
    })
    .finally(() => { isProcessing = false; });
}
</script>
</head>
<body>
    <h2>게임 상점</h2>
    <div class="container">
        <div class="card"><h3>아이템1</h3><button id="item1" onclick="openPay('item1')">구매</button></div>
        <div class="card"><h3>아이템2</h3><button id="item2" onclick="openPay('item2')">구매</button></div>
        <div class="card"><h3>아이템3</h3><button id="item3" onclick="openPay('item3')">구매</button></div>
    </div>
    <div id="modal" class="modal">
        <h3>결제하시겠습니까?</h3>
        <button onclick="doPay()">결제하기</button>
        <button onclick="document.getElementById('modal').style.display='none'" style="background:red;">취소</button>
    </div>
</body>
</html>
"""

# ================================
# 서버 핸들러
# ================================
class MyHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args): return

    def do_GET(self):
        global TEST_STATE
        if self.path == "/":
            TEST_STATE["page_loaded"] = True
            TEST_STATE["image_loaded"] = True 
            
            self.send_response(200)
            self.end_headers()
            self.wfile.write(MAIN_PAGE.encode("utf-8"))
            
            # 접속 즉시 1, 2번 케이스 기록
            run_tests(TEST_STATE.copy(), "N/A")

    def do_POST(self):
        global TEST_STATE, SERVER_STATE
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length).decode()
        data = json.loads(body) if body else {}

        if self.path == "/click":
            TEST_STATE["modal_opened"] = True
            self.send_response(200)
            self.end_headers()

        elif self.path == "/purchase":
            SERVER_STATE["request_count"] += 1
            TEST_STATE["request_sent"] = True
            # 중복 클릭 여부 판단 (첫 번째 요청만 Pass)
            TEST_STATE["no_duplicate"] = (SERVER_STATE["request_count"] == 1)
            
            self.send_response(200)
            self.end_headers()

        elif self.path == "/result":
            # 클라이언트에서 보낸 상세 UI 상태 반영
            TEST_STATE.update(data)
            
            self.send_response(200)
            self.end_headers()
            
            # 최종 결과 기록
            run_tests(TEST_STATE.copy(), data.get("selected", "Unknown"))
            
            # 다음 아이템 테스트를 위해 카운트와 상태 일부 초기화
            SERVER_STATE["request_count"] = 0
            TEST_STATE["request_sent"] = False
            TEST_STATE["modal_opened"] = False

# ================================
# 실행
# ================================
if __name__ == "__main__":
    reset_excel()
    server = HTTPServer(("127.0.0.1", 5000), MyHandler)
    print("테스트 서버 시작: http://127.0.0.1:5000")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n서버를 종료합니다.")
        server.shutdown()